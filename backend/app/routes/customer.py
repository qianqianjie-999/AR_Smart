"""客户管理路由"""

from datetime import date
from flask import Blueprint, g, request, send_file
from sqlalchemy import func

from app import db
from app.models import Customer, CustomerContact, Contract
from app.utils.auth import login_required
from app.utils.helpers import success_response, error_response, paginate, log_operation

customer_bp = Blueprint('customer', __name__, url_prefix='/api/customer')


@customer_bp.route('/', methods=['GET'])
@login_required
def list_customers():
    """客户列表（分页）"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    region = request.args.get('region', '')
    name = request.args.get('name', '')
    credit_level = request.args.get('credit_level', '')

    query = db.session.query(
        Customer.id,
        Customer.region,
        Customer.name,
        Customer.business_contact,
        Customer.credit_level,
        func.coalesce(func.sum(Contract.outstanding_amount), 0).label('outstanding_amount'),
        Customer.created_at,
    ).outerjoin(Contract, (Contract.customer_id == Customer.id) & (Contract.is_deleted == 0))

    # 排除已删除
    query = query.filter(Customer.is_deleted == 0)

    # 模糊搜索
    if name:
        query = query.filter(Customer.name.contains(name))
    if region:
        query = query.filter(Customer.region == region)
    if credit_level:
        query = query.filter(Customer.credit_level == credit_level)

    query = query.group_by(Customer.id).order_by(Customer.created_at.desc())

    result = paginate(query, page=page, page_size=page_size)

    # Serialize
    data = []
    for item in result['list']:
        data.append({
            'id': item.id,
            'region': item.region,
            'name': item.name,
            'business_contact': item.business_contact,
            'credit_level': item.credit_level,
            'outstanding_amount': float(item.outstanding_amount) if item.outstanding_amount else 0,
            'created_at': item.created_at.strftime('%Y-%m-%d %H:%M:%S') if item.created_at else None,
        })

    return success_response({
        'list': data,
        'total': result['total'],
        'page': result['page'],
        'page_size': result['page_size'],
    })


@customer_bp.route('/regions', methods=['GET'])
@login_required
def get_regions():
    """获取所有区域列表（去重）"""
    regions = db.session.query(Customer.region).filter(
        Customer.is_deleted == 0,
        Customer.region.isnot(None),
        Customer.region != '',
    ).distinct().order_by(Customer.region).all()
    region_list = [r[0] for r in regions if r[0]]
    return success_response(region_list)


@customer_bp.route('/<int:id>', methods=['GET'])
@login_required
def get_customer(id):
    """客户详情"""
    customer = Customer.query.filter_by(id=id, is_deleted=0).first()
    if not customer:
        return error_response('客户不存在', 404)

    contacts = CustomerContact.query.filter_by(customer_id=id).all()
    contacts_data = []
    for c in contacts:
        contacts_data.append({
            'id': c.id,
            'contact_type': c.contact_type,
            'name': c.name,
            'phone': c.phone,
            'email': c.email,
            'position': c.position,
            'is_primary': c.is_primary,
        })

    data = {
        'id': customer.id,
        'region': customer.region,
        'name': customer.name,
        'registered_addr': customer.registered_addr,
        'contact_addr': customer.contact_addr,
        'billing_info': customer.billing_info,
        'business_contact': customer.business_contact,
        'credit_level': customer.credit_level,
        'remark': customer.remark,
        'created_at': customer.created_at.strftime('%Y-%m-%d %H:%M:%S') if customer.created_at else None,
        'updated_at': customer.updated_at.strftime('%Y-%m-%d %H:%M:%S') if customer.updated_at else None,
        'contacts': contacts_data,
    }
    return success_response(data)


@customer_bp.route('/', methods=['POST'])
@login_required
def create_customer():
    """新增客户"""
    json_data = request.get_json(silent=True)
    if not json_data:
        return error_response('请提供有效的JSON数据')

    name = json_data.get('name', '')
    if not name:
        return error_response('客户名称不能为空')

    customer = Customer(
        region=json_data.get('region', ''),
        name=name,
        registered_addr=json_data.get('registered_addr', ''),
        contact_addr=json_data.get('contact_addr', ''),
        billing_info=json_data.get('billing_info', ''),
        business_contact=json_data.get('business_contact', ''),
        credit_level=json_data.get('credit_level', 'B'),
        remark=json_data.get('remark', ''),
        created_by=getattr(g, 'current_user_id', None),
        updated_by=getattr(g, 'current_user_id', None),
    )
    db.session.add(customer)
    db.session.flush()  # 获取 customer.id

    # 创建联系人
    contacts = json_data.get('contacts', [])
    for c in contacts:
        contact = CustomerContact(
            customer_id=customer.id,
            contact_type=c.get('contact_type', ''),
            name=c.get('name', ''),
            phone=c.get('phone', ''),
            email=c.get('email', ''),
            position=c.get('position', ''),
            is_primary=c.get('is_primary', 0),
        )
        db.session.add(contact)

    db.session.commit()

    log_operation(
        user_id=g.get('current_user_id'),
        username=g.get('current_username', ''),
        module='客户管理',
        action='新增客户',
        content=f'新增客户: {name}',
    )

    return success_response({'id': customer.id}, '新增客户成功')


@customer_bp.route('/<int:id>', methods=['PUT'])
@login_required
def update_customer(id):
    """编辑客户"""
    customer = Customer.query.filter_by(id=id, is_deleted=0).first()
    if not customer:
        return error_response('客户不存在', 404)

    json_data = request.get_json(silent=True)
    if not json_data:
        return error_response('请提供有效的JSON数据')

    customer.region = json_data.get('region', customer.region)
    customer.name = json_data.get('name', customer.name)
    customer.registered_addr = json_data.get('registered_addr', customer.registered_addr)
    customer.contact_addr = json_data.get('contact_addr', customer.contact_addr)
    customer.billing_info = json_data.get('billing_info', customer.billing_info)
    customer.business_contact = json_data.get('business_contact', customer.business_contact)
    customer.credit_level = json_data.get('credit_level', customer.credit_level)
    customer.remark = json_data.get('remark', customer.remark)
    customer.updated_by = g.get('current_user_id')

    # 更新联系人：先删后增
    contacts = json_data.get('contacts')
    if contacts is not None:
        CustomerContact.query.filter_by(customer_id=id).delete()
        for c in contacts:
            contact = CustomerContact(
                customer_id=id,
                contact_type=c.get('contact_type', ''),
                name=c.get('name', ''),
                phone=c.get('phone', ''),
                email=c.get('email', ''),
                position=c.get('position', ''),
                is_primary=c.get('is_primary', 0),
            )
            db.session.add(contact)

    db.session.commit()

    log_operation(
        user_id=g.get('current_user_id'),
        username=g.get('current_username', ''),
        module='客户管理',
        action='编辑客户',
        content=f'编辑客户: {customer.name}',
    )

    return success_response(message='编辑客户成功')


@customer_bp.route('/<int:id>', methods=['DELETE'])
@login_required
def delete_customer(id):
    """删除客户（软删除）"""
    customer = Customer.query.filter_by(id=id, is_deleted=0).first()
    if not customer:
        return error_response('客户不存在', 404)

    customer.is_deleted = 1
    customer.updated_by = g.get('current_user_id')
    db.session.commit()

    log_operation(
        user_id=g.get('current_user_id'),
        username=g.get('current_username', ''),
        module='客户管理',
        action='删除客户',
        content=f'删除客户: {customer.name}',
    )

    return success_response(message='删除客户成功')


@customer_bp.route('/export', methods=['GET'])
@login_required
def export_customers():
    """导出客户Excel"""
    try:
        import openpyxl
        from io import BytesIO
    except ImportError:
        return error_response('导出功能需要安装 openpyxl 库', 500)

    customers = Customer.query.filter_by(is_deleted=0).order_by(Customer.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '客户列表'
    headers = ['ID', '区域', '客户名称', '注册地址', '联系地址', '开票信息', '业务联系人', '信用等级', '备注', '创建时间']
    ws.append(headers)

    for c in customers:
        ws.append([
            c.id,
            c.region,
            c.name,
            c.registered_addr,
            c.contact_addr,
            c.billing_info,
            c.business_contact,
            c.credit_level,
            c.remark,
            c.created_at.strftime('%Y-%m-%d %H:%M:%S') if c.created_at else '',
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'客户列表_{date.today().strftime("%Y%m%d")}.xlsx',
    )


@customer_bp.route('/import', methods=['POST'])
@login_required
def import_customers():
    """批量导入客户（Excel）"""
    try:
        import openpyxl
    except ImportError:
        return error_response('导入功能需要安装 openpyxl 库', 500)

    if 'file' not in request.files:
        return error_response('请上传Excel文件')

    file = request.files['file']
    if file.filename == '':
        return error_response('请选择文件')

    ext = file.filename.rsplit('.', 1)[-1].lower()
    if ext not in ('xlsx', 'xls'):
        return error_response('仅支持 .xlsx / .xls 格式', 400)

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
    except Exception as e:
        return error_response(f'Excel读取失败: {str(e)}', 400)

    # 读取表头（第一行）
    headers = [cell.value for cell in ws[1]]
    if not headers:
        return error_response('Excel文件为空或格式不正确', 400)

    # 列名映射（支持中文列名）
    column_map = {
        '区域': 'region', '客户名称': 'name', '注册地址': 'registered_addr',
        '联系地址': 'contact_addr', '开票信息': 'billing_info',
        '业务联系人': 'business_contact', '信用等级': 'credit_level', '备注': 'remark',
        '联系人姓名': 'contact_name', '联系人职位': 'contact_position',
        '联系人电话': 'contact_phone', '联系人邮箱': 'contact_email',
    }

    # 建立列索引映射
    col_index = {}
    for i, h in enumerate(headers):
        if h and h.strip() in column_map:
            col_index[column_map[h.strip()]] = i

    if 'name' not in col_index:
        return error_response('Excel缺少"客户名称"列', 400)

    success_count = 0
    error_rows = []
    current_customer = None
    current_customer_name = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            name = str(row[col_index['name']]).strip() if col_index['name'] < len(row) and row[col_index['name']] else ''

            # 如果客户名称变化，创建新客户
            if name and name != current_customer_name:
                customer = Customer(
                    region=str(row[col_index.get('region', -1)]).strip() if col_index.get('region', -1) >= 0 and col_index['region'] < len(row) and row[col_index['region']] else '',
                    name=name,
                    registered_addr=str(row[col_index.get('registered_addr', -1)]).strip() if col_index.get('registered_addr', -1) >= 0 and col_index['registered_addr'] < len(row) and row[col_index['registered_addr']] else '',
                    contact_addr=str(row[col_index.get('contact_addr', -1)]).strip() if col_index.get('contact_addr', -1) >= 0 and col_index['contact_addr'] < len(row) and row[col_index['contact_addr']] else '',
                    billing_info=str(row[col_index.get('billing_info', -1)]).strip() if col_index.get('billing_info', -1) >= 0 and col_index['billing_info'] < len(row) and row[col_index['billing_info']] else '',
                    business_contact=str(row[col_index.get('business_contact', -1)]).strip() if col_index.get('business_contact', -1) >= 0 and col_index['business_contact'] < len(row) and row[col_index['business_contact']] else '',
                    credit_level=str(row[col_index.get('credit_level', -1)]).strip() if col_index.get('credit_level', -1) >= 0 and col_index['credit_level'] < len(row) and row[col_index['credit_level']] else 'B',
                    remark=str(row[col_index.get('remark', -1)]).strip() if col_index.get('remark', -1) >= 0 and col_index['remark'] < len(row) and row[col_index['remark']] else '',
                    created_by=getattr(g, 'current_user_id', None),
                )
                db.session.add(customer)
                db.session.flush()
                current_customer = customer
                current_customer_name = name
                success_count += 1

            # 添加联系人
            contact_name = str(row[col_index.get('contact_name', -1)]).strip() if col_index.get('contact_name', -1) >= 0 and col_index['contact_name'] < len(row) and row[col_index['contact_name']] else ''
            if contact_name and current_customer:
                contact = CustomerContact(
                    customer_id=current_customer.id,
                    contact_type='商务',
                    name=contact_name,
                    phone=str(row[col_index.get('contact_phone', -1)]).strip() if col_index.get('contact_phone', -1) >= 0 and col_index['contact_phone'] < len(row) and row[col_index['contact_phone']] else '',
                    email=str(row[col_index.get('contact_email', -1)]).strip() if col_index.get('contact_email', -1) >= 0 and col_index['contact_email'] < len(row) and row[col_index['contact_email']] else '',
                    position=str(row[col_index.get('contact_position', -1)]).strip() if col_index.get('contact_position', -1) >= 0 and col_index['contact_position'] < len(row) and row[col_index['contact_position']] else '',
                )
                db.session.add(contact)

        except Exception as e:
            error_rows.append(f'第{row_idx}行: {str(e)}')

    db.session.commit()

    log_operation(
        user_id=getattr(g, 'current_user_id', None),
        username=getattr(g, 'current_username', ''),
        module='客户管理',
        action='批量导入',
        content=f'批量导入客户 {success_count} 条',
    )

    return success_response({
        'success_count': success_count,
        'error_rows': error_rows,
    }, f'导入完成：成功 {success_count} 条' + (f'，失败 {len(error_rows)} 条' if error_rows else ''))
